import csv
from pathlib import Path
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]

NOTIFICATION_TEST_DATA_FILE = (
    PROJECT_ROOT
    / "TestCases"
    / "Notification_TestCase.xlsx"
)

NOTIFICATION_TEST_DATA_CSV = (
    Path(__file__).resolve().parents[1]
    / "test_data"
    / "notification_test_data.csv"
)


def get_test_data(file_path, test_case_id):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook["AUTOMATION_DATA"]

    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))

        if row_data["test_case_id"] == test_case_id:
            workbook.close()
            return row_data

    workbook.close()
    raise ValueError(
        f"Không tìm thấy test case: {test_case_id}"
    )


def get_test_data_csv(file_path, test_case_id):
    with open(file_path, mode="r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)

        for row in reader:
            if row["test_case_id"] == test_case_id:
                return row

    raise ValueError(
        f"Không tìm thấy test case trong CSV: {test_case_id}"
    )